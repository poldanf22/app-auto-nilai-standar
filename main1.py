import streamlit as st
import openpyxl
import pandas as pd
from openpyxl.styles import Font
from io import StringIO

st.title('Olah Nilai Standar')


uploaded_file = st.file_uploader('Letakkan file excel', type='xlsx')
if uploaded_file is not None:
	dataframe = pd.read_excel(uploaded_file)
	st.write(dataframe)

	wb = openpyxl.load_workbook(uploaded_file)
	ws = wb['nilai_std_pts']

	q=len(ws['K'])
	r=len(ws['K'])+2
	s=len(ws['K'])+3


	ws['G{}'.format(r)] = "=AVERAGE(G2:G{})".format(q)
	ws['H{}'.format(r)] = "=AVERAGE(H2:H{})".format(q)
	ws['I{}'.format(r)] = "=AVERAGE(I2:I{})".format(q)
	ws['J{}'.format(r)] = "=AVERAGE(J2:J{})".format(q)
	ws['K{}'.format(r)] = "=AVERAGE(K2:K{})".format(q)
	ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
	ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
	ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
	ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
	ws['K{}'.format(s)] = "=STDEV(K2:K{})".format(q)
	ws['L{}'.format(r)] = "=MAX(L2:L{})".format(q)
	ws['M{}'.format(r)] = "=MAX(M2:M{})".format(q)
	ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
	ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
	ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)

	# Z Score
	ws['L1'] = 'Z_MAT'
	ws['M1'] = 'Z_IND'
	ws['N1'] = 'Z_ENG'
	ws['O1'] = 'Z_IPA'
	ws['P1'] = 'Z_IPS'
	ws['Q1'] = 'S_MAT'
	ws['R1'] = 'S_IND'
	ws['S1'] = 'S_ENG'
	ws['T1'] = 'S_IPA'
	ws['U1'] = 'S_IPS'
	ws['V1'] = 'S_JML'
	ws['W1'] = 'RANK'
	ws['L1'].font = Font(bold=False, name='Calibri', size=11)
	ws['M1'].font = Font(bold=False, name='Calibri', size=11)
	ws['N1'].font = Font(bold=False, name='Calibri', size=11)
	ws['O1'].font = Font(bold=False, name='Calibri', size=11)
	ws['P1'].font = Font(bold=False, name='Calibri', size=11)
	ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
	ws['R1'].font = Font(bold=False, name='Calibri', size=11)
	ws['S1'].font = Font(bold=False, name='Calibri', size=11)
	ws['T1'].font = Font(bold=False, name='Calibri', size=11)
	ws['U1'].font = Font(bold=False, name='Calibri', size=11)
	ws['V1'].font = Font(bold=False, name='Calibri', size=11)
	ws['W1'].font = Font(bold=False, name='Calibri', size=11)

	for row in range (2, q+1):
		ws['L{}'.format(row)]='=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row,row,r,s)
		ws['M{}'.format(row)]='=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row,row,r,s)
		ws['N{}'.format(row)]='=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row,row,r,s)
		ws['O{}'.format(row)]='=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row,row,r,s)
		ws['P{}'.format(row)]='=IFERROR(ROUND(IF(K{}="","",(K{}-K${})/K${}),2),"")'.format(row,row,r,s)
		ws['Q{}'.format(row)]='=IFERROR(ROUND(IF(G{}="","",IF(70+30*L{}/$L${}<20,20,70+30*L{}/$L${})),2),"")'.format(row,row,r,row,r)
		ws['R{}'.format(row)]='=IFERROR(ROUND(IF(H{}="","",IF(70+30*M{}/$M${}<20,20,70+30*M{}/$M${})),2),"")'.format(row,row,r,row,r)
		ws['S{}'.format(row)]='=IFERROR(ROUND(IF(I{}="","",IF(70+30*N{}/$N${}<20,20,70+30*N{}/$N${})),2),"")'.format(row,row,r,row,r)
		ws['T{}'.format(row)]='=IFERROR(ROUND(IF(J{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$O${})),2),"")'.format(row,row,r,row,r)
		ws['U{}'.format(row)]='=IFERROR(ROUND(IF(K{}="","",IF(70+30*P{}/$P${}<20,20,70+30*P{}/$P${})),2),"")'.format(row,row,r,row,r)

		ws['V{}'.format(row)]='=IF(SUM(Q{}:U{})=0,"",SUM(Q{}:U{}))'.format(row,row,row,row)
		ws['W{}'.format(row)]='=IF(V{}="","",RANK(V{},$V$2:$V${}))'.format(row,row,q)

	wb.save(r"C:Desktop\nilai_std_pts_dummy.xlsx")	

